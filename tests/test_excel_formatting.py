import pandas as pd
import openpyxl
from pathlib import Path
from helpers.excel_formatting import save_formatted_excel


def test_value_format_percent_and_integer(tmp_path):
    df = pd.DataFrame({
        "Name": ["A", "B"],
        "Pct": [0.12345, 0.5],
        "Count": [1, 2],
    })
    out_path = tmp_path / "fmt_test.xlsx"
    # Request percent formatting on Pct and integer on Count
    save_formatted_excel(df, out_path, value_formats={"Pct": "percent", "Count": "int"}, autofilter=False)
    # Read workbook with openpyxl
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    # First data row is row 2 (header row 1)
    pct_cell = ws.cell(row=2, column=2)  # B2
    count_cell = ws.cell(row=2, column=3)  # C2
    assert pct_cell.number_format == "0.#####%", f"Unexpected number_format: {pct_cell.number_format}"
    assert count_cell.number_format in ("0", "#,#0", "0_ ") or count_cell.number_format.startswith("0"), f"Unexpected number_format: {count_cell.number_format}"
