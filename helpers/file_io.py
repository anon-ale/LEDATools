
from pathlib import Path
from typing import Optional, List
import pandas as pd
from openpyxl import load_workbook
from helpers.constants import EXCEL_EXTENSIONS

from PySide6.QtWidgets import QFileDialog, QMessageBox, QWidget

def read_data_file(path, read_all_sheets=False):
    """Read data from a file (CSV or Excel).
    
    Parameters:
    - path: file path (str or Path)
    - read_all_sheets: if True and the file is Excel, return a dict of {sheet_name: DataFrame}.
                       if False (default), return a single DataFrame from the first sheet (CSV or Excel).
    
    Returns:
    - Single DataFrame (if read_all_sheets=False)
    - Dict[str, DataFrame] (if read_all_sheets=True and file is Excel)
    - None (if file cannot be read)
    """
    p = Path(path)

    # CSV handling goes here...
    if p.suffix.lower() == ".csv":
        if read_all_sheets:
            # CSV doesn't have multiple sheets; return a dict with one entry
            try:
                for enc in ["utf-8", "latin1", "windows-1252"]:
                    try:
                        df = pd.read_csv(p, encoding=enc)
                        return {p.stem: df}
                    except Exception:
                        continue
                df = pd.read_csv(p, engine="python", on_bad_lines="skip")
                return {p.stem: df}
            except Exception as e:
                print(f"Skipped unreadable CSV file: {p.name} ({e})")
                return None
        else:
            for enc in ["utf-8", "latin1", "windows-1252"]:
                try:
                    return pd.read_csv(p, encoding=enc)
                except Exception:
                    continue
            return pd.read_csv(p, engine="python", on_bad_lines="skip")

    # Excel handling using known Excel extensions, with a robust fallback
    suffix = p.suffix.lower()
    try:
        # If the suffix explicitly indicates Excel, use pandas.read_excel directly
        if suffix in EXCEL_EXTENSIONS:
            if read_all_sheets:
                return pd.read_excel(p, sheet_name=None)
            return pd.read_excel(p)

        # Unknown extension: attempt to read as Excel first (pandas supports many formats)
        if read_all_sheets:
            return pd.read_excel(p, sheet_name=None)
        return pd.read_excel(p)
    except Exception:
        # Fallback: read values only, ignore broken XML/styles using openpyxl
        try:
            wb = load_workbook(filename=p, data_only=True)
            if read_all_sheets:
                result = {}
                for sheet_name in wb.sheetnames:
                    sheet = wb[sheet_name]
                    data = list(sheet.values)
                    if not data:
                        continue
                    cols = data[0]
                    result[sheet_name] = pd.DataFrame(data[1:], columns=cols)
                return result if result else None
            else:
                sheet = wb[wb.sheetnames[0]]
                data = sheet.values
                cols = next(data)  # first row as header
                return pd.DataFrame(data, columns=cols)
        except Exception as e:
            print(f"Skipped unreadable Excel file: {p.name} ({e})")
            return None

def ask_for_file(parent: QWidget, caption: str = "Select File") -> Optional[str]:
    file_path, _ = QFileDialog.getOpenFileName(
        parent,
        caption,
        "",
        "Data Files (*.xlsx *.xls *.csv)"
    )
    return file_path or None


def ask_for_multiple_files(parent: QWidget, caption: str = "Select File(s)") -> List[str]:
    files, _ = QFileDialog.getOpenFileNames(
        parent,
        caption,
        "",
        "Data Files (*.xlsx *.xls *.csv)"
    )
    return files or []


def ask_for_save_excel(parent: QWidget, caption: str = "Save Output As") -> Optional[str]:
    file_path, _ = QFileDialog.getSaveFileName(
        parent,
        caption,
        "",
        "Excel Files (*.xlsx)"
    )
    if not file_path:
        return None

    path = Path(file_path)
    if path.suffix.lower() != ".xlsx":
        path = path.with_suffix(".xlsx")
    return str(path)


def show_info(parent: QWidget, title: str, message: str) -> None:
    QMessageBox.information(parent, title, message)


def show_error(parent: QWidget, title: str, message: str) -> None:
    QMessageBox.critical(parent, title, message)


def show_warning(parent: QWidget, title: str, message: str) -> None:
    QMessageBox.warning(parent, title, message)
